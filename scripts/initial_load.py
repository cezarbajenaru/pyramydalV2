#!/usr/bin/env python3
"""
Initial data load script: Load existing Excel data into PostgreSQL
"""

import argparse
import sys
from datetime import datetime, timedelta
import psycopg2
import openpyxl
from typing import Dict, Any


def normalize_excel_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().replace("\n", " ")


def parse_args():
    parser = argparse.ArgumentParser(description='Load Excel data into PostgreSQL')
    parser.add_argument('--excel', required=True, help='Path to Excel file')
    parser.add_argument('--sheet', default='PAGINA PRINCIPALA', help='Sheet name')
    parser.add_argument('--db-host', required=True, help='Database host')
    parser.add_argument('--db-port', type=int, default=5432, help='Database port')
    parser.add_argument('--db-name', default='production', help='Database name')
    parser.add_argument('--db-user', required=True, help='Database user')
    parser.add_argument('--db-password', required=True, help='Database password')
    parser.add_argument('--dry-run', action='store_true', help='Dry run (no commits)')
    parser.add_argument('--max-rows', type=int, default=0, help='Limit rows for test run (0 = all)')
    return parser.parse_args()


def connect_db(host, port, dbname, user, password):
    """Connect to PostgreSQL"""
    print(f"Connecting to PostgreSQL: {host}/{dbname}")
    return psycopg2.connect(
        host=host,
        port=port,
        database=dbname,
        user=user,
        password=password
    )


def parse_excel(filepath, sheet_name):
    """Parse Excel file"""
    print(f"Loading Excel file: {filepath}")
    print(f"Sheet: {sheet_name}")
    
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    
    if sheet_name not in wb.sheetnames:
        print(f"❌ Sheet '{sheet_name}' not found!")
        print(f"Available sheets: {wb.sheetnames}")
        sys.exit(1)
    
    ws = wb[sheet_name]
    
    print(f"Dimensions: {ws.max_row} rows x {ws.max_column} cols")
    
    # Get headers
    headers_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    headers = [str(h).strip() if h else f'col_{i}' for i, h in enumerate(headers_row)]
    
    print(f"Headers: {headers[:10]}...")
    
    # Column mapping (adjust based on your Excel structure)
    column_mapping = {
        'nr fisa': 'nr_fisa',
        'reper': 'reper',
        'client': 'client',
        'buc.': 'buc',
        'timp//buc': 'timp_per_buc',
        'strung (colchester + cazeneuve + tos)': 'strung_colchester',
        'strung cnc (sbl 500 + talent 51)': 'strung_cnc',
        'freze mici (schaublin 53, schaublin 53n)': 'freze_mici',
        'freze mari (shw nc + gambin)': 'freze_mari',
        'gaurire (shw nc + aciera)': 'gaurire',
        'rectificare': 'rectificare',
        'bwk': 'bwk',
        'sip': 'sip',
        'norte': 'norte',
        'tos': 'tos',
        'bridgeport': 'bridgeport',
        'eco': 'eco',
        'schaublin': 'schaublin',
        'hurco': 'hurco',
        'matec': 'matec',
        'parpas': 'parpas',
        'ajustare': 'ajustare',
        'filetare': 'filetare',
        'marcare': 'marcare',
        'curatare filete': 'curatare_filete',
        'program': 'locatie_dosar',
        'programator': 'programator',
        'obs': 'observatii',
        'termen livrare': 'data_livrare',
    }
    
    # Parse data rows
    data = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):  # Skip empty rows
            continue
        
        row_dict = {}
        for header, cell_value in zip(headers, row):
            normalized_header = normalize_excel_header(header)
            db_col = column_mapping.get(normalized_header, normalized_header.replace(' ', '_'))
            row_dict[db_col] = cell_value
        
        row_dict['_source_row'] = row_idx
        data.append(row_dict)
    
    wb.close()
    
    print(f"Parsed {len(data)} data rows")
    return data


def clean_data(data):
    """Clean and validate data"""
    print("Cleaning data...")
    
    cleaned = []
    errors = []
    
    for row in data:
        try:
            # Trim strings
            if 'reper' in row and row['reper']:
                row['reper'] = str(row['reper']).strip()
            if 'client' in row and row['client']:
                row['client'] = str(row['client']).strip()
            
            # Validate required fields
            if not row.get('nr_fisa'):
                errors.append(f"Row {row['_source_row']}: Missing nr_fisa")
                continue
            if not row.get('reper'):
                errors.append(f"Row {row['_source_row']}: Missing reper")
                continue
            
            cleaned.append(row)
        except Exception as e:
            errors.append(f"Row {row['_source_row']}: {str(e)}")
    
    if errors:
        print(f"⚠️  {len(errors)} rows with errors:")
        for err in errors[:10]:  # Show first 10
            print(f"  - {err}")
    
    print(f"✅ {len(cleaned)} rows cleaned successfully")
    return cleaned, errors


def normalize_date_value(value):
    """Convert Excel date serials/strings to Python date when possible."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
        return value
    # Excel date serial (days since 1899-12-30)
    if isinstance(value, (int, float)):
        if value <= 0:
            return None
        excel_epoch = datetime(1899, 12, 30)
        return (excel_epoch + timedelta(days=float(value))).date()
    return None


def normalize_numeric_value(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if text == "":
        return None
    text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def normalize_buc_value(value):
    buc = normalize_numeric_value(value)
    if buc is None:
        return None
    if buc <= 0:
        return None
    return int(round(buc))


def load_to_db(db_conn, data, dry_run=False):
    """Load data into PostgreSQL"""
    print(f"Loading {len(data)} rows to database...")
    
    if dry_run:
        print("🔍 DRY RUN - No data will be committed")
    
    cursor = db_conn.cursor()
    
    inserted = 0
    failed = 0
    
    for row in data:
        try:
            cursor.execute("""
                INSERT INTO app.main_rows (
                    nr_fisa, reper, client, buc,
                    timp_per_buc,
                    strung_colchester, strung_cnc, freze_mici, freze_mari, gaurire, rectificare,
                    bwk, sip, norte, tos, bridgeport, eco, schaublin, hurco, matec, parpas,
                    ajustare, filetare, marcare, curatare_filete,
                    locatie_dosar, programator, observatii, data_livrare,
                    created_by, created_at
                ) VALUES (
                    %s, %s, %s, %s,
                    %s,
                    %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s
                )
            """, (
                row.get('nr_fisa'),
                row.get('reper'),
                row.get('client'),
                normalize_buc_value(row.get('buc')),
                normalize_numeric_value(row.get('timp_per_buc')),
                normalize_numeric_value(row.get('strung_colchester')),
                normalize_numeric_value(row.get('strung_cnc')),
                normalize_numeric_value(row.get('freze_mici')),
                normalize_numeric_value(row.get('freze_mari')),
                normalize_numeric_value(row.get('gaurire')),
                normalize_numeric_value(row.get('rectificare')),
                normalize_numeric_value(row.get('bwk')),
                normalize_numeric_value(row.get('sip')),
                normalize_numeric_value(row.get('norte')),
                normalize_numeric_value(row.get('tos')),
                normalize_numeric_value(row.get('bridgeport')),
                normalize_numeric_value(row.get('eco')),
                normalize_numeric_value(row.get('schaublin')),
                normalize_numeric_value(row.get('hurco')),
                normalize_numeric_value(row.get('matec')),
                normalize_numeric_value(row.get('parpas')),
                normalize_numeric_value(row.get('ajustare')),
                normalize_numeric_value(row.get('filetare')),
                normalize_numeric_value(row.get('marcare')),
                normalize_numeric_value(row.get('curatare_filete')),
                row.get('locatie_dosar'),
                row.get('programator'),
                row.get('observatii'),
                normalize_date_value(row.get('data_livrare')),
                'migration',
                datetime.now()
            ))
            if dry_run:
                db_conn.rollback()
            else:
                db_conn.commit()
            inserted += 1
        except Exception as e:
            db_conn.rollback()
            print(f"❌ Failed to insert row {row.get('_source_row')}: {e}")
            failed += 1
    
    if dry_run:
        print("🔄 Dry run checked each row without persisting")
    else:
        print("✅ Committed to database")
    
    cursor.close()
    
    print(f"Summary: {inserted} inserted, {failed} failed")
    return inserted, failed


def main():
    args = parse_args()
    
    print("=" * 50)
    print("PyramydalV2 Initial Data Load")
    print("=" * 50)
    
    # Parse Excel
    data = parse_excel(args.excel, args.sheet)
    
    # Clean data
    cleaned_data, errors = clean_data(data)
    
    if not cleaned_data:
        print("❌ No valid data to load!")
        sys.exit(1)
    
    # Connect to DB
    db_conn = connect_db(args.db_host, args.db_port, args.db_name, args.db_user, args.db_password)
    
    # Optional row limit for faster validation
    data_to_load = cleaned_data if args.max_rows <= 0 else cleaned_data[:args.max_rows]
    if args.max_rows > 0:
        print(f"Limiting load to first {len(data_to_load)} rows")

    # Load data
    inserted, failed = load_to_db(db_conn, data_to_load, dry_run=args.dry_run)
    
    # Close connection
    db_conn.close()
    
    print("=" * 50)
    if args.dry_run:
        print("✅ Dry run completed successfully!")
        print("Run without --dry-run to commit data")
    else:
        print("✅ Data load completed!")
        print(f"Inserted: {inserted} rows")
        if failed > 0:
            print(f"⚠️  Failed: {failed} rows")
    print("=" * 50)


if __name__ == '__main__':
    main()

