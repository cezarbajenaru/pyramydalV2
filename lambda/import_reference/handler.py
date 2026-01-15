"""
Lambda function: Import Reference Data
Handles XLSX uploads from S3, validates, and loads into PostgreSQL staging tables
"""

import json
import os
import uuid
from datetime import datetime
from typing import Dict, List, Any, Optional
import boto3
import psycopg2
import openpyxl
from io import BytesIO

# Environment variables
DB_HOST = os.environ['DB_HOST']
DB_NAME = os.environ['DB_NAME']
DB_USER = os.environ['DB_USER']
DB_PASSWORD = os.environ['DB_PASSWORD']
S3_BUCKET = os.environ['S3_BUCKET']

s3_client = boto3.client('s3')


def lambda_handler(event, context):
    """
    Main handler for import operations
    
    Event structure:
    {
        "s3_key": "uploads/lista_programe/file.xlsx",
        "import_type": "lista_programe",  # or "price_list", "timing_list", "cnc_times"
        "uploaded_by": "user@example.com",
        "sheet_name": "Sheet1"  # optional, defaults to first sheet
    }
    """
    
    try:
        # Parse input
        s3_key = event.get('s3_key')
        import_type = event.get('import_type')
        uploaded_by = event.get('uploaded_by', 'unknown')
        sheet_name = event.get('sheet_name')
        
        if not s3_key or not import_type:
            return error_response('Missing required parameters: s3_key and import_type')
        
        # Generate import ID
        import_id = str(uuid.uuid4())
        
        print(f"[{import_id}] Starting import: {import_type} from {s3_key}")
        
        # Initialize audit record
        db_conn = get_db_connection()
        init_audit_record(db_conn, import_id, import_type, s3_key, uploaded_by, context.request_id)
        
        # Download XLSX from S3
        print(f"[{import_id}] Downloading from S3...")
        xlsx_bytes = download_from_s3(s3_key)
        
        # Parse XLSX
        print(f"[{import_id}] Parsing XLSX...")
        data = parse_xlsx(xlsx_bytes, import_type, sheet_name)
        
        # Validate data
        print(f"[{import_id}] Validating data...")
        validation_errors = validate_data(data, import_type)
        
        if validation_errors:
            print(f"[{import_id}] Validation failed: {len(validation_errors)} errors")
            update_audit_failed(db_conn, import_id, validation_errors, len(data))
            db_conn.close()
            return error_response(f"Validation failed with {len(validation_errors)} errors", validation_errors)
        
        # Load into staging table
        print(f"[{import_id}] Loading into staging table...")
        rows_loaded = load_to_staging(db_conn, data, import_type)
        
        # Run SQL validation
        print(f"[{import_id}] Running SQL validation...")
        sql_errors = run_sql_validation(db_conn, import_type)
        
        if sql_errors:
            print(f"[{import_id}] SQL validation failed")
            update_audit_failed(db_conn, import_id, sql_errors, rows_loaded)
            db_conn.close()
            return error_response("SQL validation failed", sql_errors)
        
        # Swap staging to live
        print(f"[{import_id}] Swapping staging to live...")
        swap_to_live(db_conn, import_type, import_id)
        
        # Success
        update_audit_success(db_conn, import_id, rows_loaded)
        db_conn.commit()
        db_conn.close()
        
        print(f"[{import_id}] Import completed successfully: {rows_loaded} rows")
        
        return {
            'statusCode': 200,
            'body': json.dumps({
                'success': True,
                'import_id': import_id,
                'rows_loaded': rows_loaded,
                'import_type': import_type
            })
        }
        
    except Exception as e:
        print(f"ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        
        return error_response(f"Import failed: {str(e)}")


def get_db_connection():
    """Establish PostgreSQL connection"""
    return psycopg2.connect(
        host=DB_HOST,
        database=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD,
        connect_timeout=10
    )


def download_from_s3(s3_key: str) -> bytes:
    """Download file from S3"""
    response = s3_client.get_object(Bucket=S3_BUCKET, Key=s3_key)
    return response['Body'].read()


def parse_xlsx(xlsx_bytes: bytes, import_type: str, sheet_name: Optional[str] = None) -> List[Dict]:
    """
    Parse XLSX file into list of dictionaries
    
    Expected structures:
    - lista_programe: CLIENT, REPER, INDICE, SOFT FOLOSIT, UTILAJ, DATA, PROGRAMATOR, OBSERVATII, LOCATIE DOSAR, TIMPI MASINARE
    - price_list: REPER, CLIENT (optional), PRET_PER_BUC, MONEDA, VALABIL_DE_LA, VALABIL_PANA_LA
    - timing_list: OPERATIE, REPER (optional), TIMP_STANDARD, DESCRIERE
    - cnc_times: MASINA, REPER, TIMP_CICLU, TIMP_SETUP, OBSERVATII
    """
    
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    
    # Select sheet
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    
    print(f"Parsing sheet: {ws.title}, dimensions: {ws.max_row} x {ws.max_column}")
    
    # Extract headers from first row
    headers_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [normalize_header(h) for h in headers_row if h]
    
    print(f"Headers found: {headers}")
    
    # Validate required headers based on import type
    required_headers = get_required_headers(import_type)
    missing_headers = set(required_headers) - set(headers)
    if missing_headers:
        raise ValueError(f"Missing required columns: {missing_headers}")
    
    # Parse data rows
    data = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):  # Skip empty rows
            continue
        
        row_dict = {}
        for header, cell_value in zip(headers, row):
            if header:
                row_dict[header] = cell_value
        
        row_dict['_source_row'] = row_idx  # For error reporting
        data.append(row_dict)
    
    wb.close()
    
    print(f"Parsed {len(data)} data rows")
    return data


def normalize_header(header: str) -> str:
    """Normalize header names to match database columns"""
    if not header:
        return ''
    
    # Remove newlines, extra spaces
    normalized = str(header).strip().replace('\n', ' ').replace('\r', ' ')
    
    # Mapping of common Excel headers to DB column names
    mappings = {
        'CLIENT': 'client',
        'REPER': 'reper',
        'INDICE': 'indice',
        'SOFT FOLOSIT': 'soft_folosit',
        'UTILAJ': 'utilaj',
        'DATA': 'data_programare',
        'PROGRAMATOR': 'programator',
        'OBSERVATII': 'observatii',
        'LOCATIE DOSAR': 'locatie_dosar',
        'TIMPI MASINARE': 'timpi_masinare',
        'PRET_PER_BUC': 'pret_per_buc',
        'MONEDA': 'moneda',
        'VALABIL_DE_LA': 'valabil_de_la',
        'VALABIL_PANA_LA': 'valabil_pana_la',
        'OPERATIE': 'operatie',
        'TIMP_STANDARD': 'timp_standard',
        'DESCRIERE': 'descriere',
        'MASINA': 'masina',
        'TIMP_CICLU': 'timp_ciclu',
        'TIMP_SETUP': 'timp_setup',
    }
    
    return mappings.get(normalized.upper(), normalized.lower().replace(' ', '_'))


def get_required_headers(import_type: str) -> List[str]:
    """Get required headers for each import type"""
    requirements = {
        'lista_programe': ['client', 'reper'],
        'price_list': ['reper', 'pret_per_buc'],
        'timing_list': ['operatie', 'timp_standard'],
        'cnc_times': ['masina', 'reper', 'timp_ciclu']
    }
    return requirements.get(import_type, [])


def validate_data(data: List[Dict], import_type: str) -> List[Dict]:
    """Validate parsed data before loading"""
    errors = []
    
    # Check for duplicate keys
    keys_seen = set()
    
    for idx, row in enumerate(data):
        row_errors = []
        
        # Type-specific validation
        if import_type == 'lista_programe':
            if not row.get('reper'):
                row_errors.append('Missing required field: reper')
            if not row.get('client'):
                row_errors.append('Missing required field: client')
            
            # Check duplicates
            key = (row.get('reper'), row.get('client'), row.get('indice', '-'))
            if key in keys_seen:
                row_errors.append(f'Duplicate key: reper={row.get("reper")}, client={row.get("client")}, indice={row.get("indice")}')
            keys_seen.add(key)
        
        elif import_type == 'price_list':
            if not row.get('reper'):
                row_errors.append('Missing required field: reper')
            if not row.get('pret_per_buc'):
                row_errors.append('Missing required field: pret_per_buc')
            
            # Validate price is positive
            try:
                price = float(row.get('pret_per_buc', 0))
                if price <= 0:
                    row_errors.append('Price must be positive')
            except (ValueError, TypeError):
                row_errors.append('Invalid price value')
        
        if row_errors:
            errors.append({
                'row': row.get('_source_row', idx),
                'errors': row_errors
            })
    
    # Limit errors reported (first 50)
    return errors[:50]


def run_sql_validation(db_conn, import_type: str) -> Optional[Dict]:
    """Run SQL-based validation on staging table"""
    if import_type == 'lista_programe':
        cursor = db_conn.cursor()
        cursor.execute("SELECT app.validate_staging_lista_programe();")
        result = cursor.fetchone()[0]
        cursor.close()
        
        if result and len(result) > 0:
            return result
    
    return None


def load_to_staging(db_conn, data: List[Dict], import_type: str) -> int:
    """Load data into staging table"""
    
    staging_table = f"app.staging_{import_type}"
    
    cursor = db_conn.cursor()
    
    # Truncate staging table
    cursor.execute(f"TRUNCATE TABLE {staging_table}")
    
    # Prepare insert statement based on import type
    if import_type == 'lista_programe':
        insert_sql = f"""
            INSERT INTO {staging_table} 
            (reper, client, indice, soft_folosit, utilaj, data_programare, 
             programator, observatii, locatie_dosar, timpi_masinare)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        
        for row in data:
            cursor.execute(insert_sql, (
                row.get('reper'),
                row.get('client'),
                row.get('indice', '-'),
                row.get('soft_folosit'),
                row.get('utilaj'),
                row.get('data_programare'),
                row.get('programator'),
                row.get('observatii'),
                row.get('locatie_dosar'),
                row.get('timpi_masinare')
            ))
    
    elif import_type == 'price_list':
        insert_sql = f"""
            INSERT INTO {staging_table}
            (reper, client, pret_per_buc, moneda, valabil_de_la, valabil_pana_la)
            VALUES (%s, %s, %s, %s, %s, %s)
        """
        
        for row in data:
            cursor.execute(insert_sql, (
                row.get('reper'),
                row.get('client'),
                row.get('pret_per_buc'),
                row.get('moneda', 'EUR'),
                row.get('valabil_de_la'),
                row.get('valabil_pana_la')
            ))
    
    db_conn.commit()
    cursor.close()
    
    return len(data)


def swap_to_live(db_conn, import_type: str, import_id: str):
    """Atomically swap staging to live table"""
    cursor = db_conn.cursor()
    cursor.execute(
        "CALL app.swap_staging_to_live(%s, %s)",
        (import_type, import_id)
    )
    db_conn.commit()
    cursor.close()


def init_audit_record(db_conn, import_id: str, import_type: str, s3_key: str, uploaded_by: str, request_id: str):
    """Initialize audit record"""
    cursor = db_conn.cursor()
    cursor.execute("""
        INSERT INTO app.imports_audit 
        (import_id, import_type, file_name, file_s3_key, file_uploaded_by, status, lambda_request_id)
        VALUES (%s, %s, %s, %s, %s, 'pending', %s)
    """, (import_id, import_type, s3_key.split('/')[-1], s3_key, uploaded_by, request_id))
    db_conn.commit()
    cursor.close()


def update_audit_success(db_conn, import_id: str, rows_loaded: int):
    """Update audit record on success"""
    cursor = db_conn.cursor()
    cursor.execute("""
        UPDATE app.imports_audit
        SET status = 'success', rows_loaded = %s, completed_at = CURRENT_TIMESTAMP
        WHERE import_id = %s
    """, (rows_loaded, import_id))
    db_conn.commit()
    cursor.close()


def update_audit_failed(db_conn, import_id: str, errors: Any, rows_rejected: int):
    """Update audit record on failure"""
    cursor = db_conn.cursor()
    cursor.execute("""
        UPDATE app.imports_audit
        SET status = 'failed', validation_errors = %s, rows_rejected = %s, completed_at = CURRENT_TIMESTAMP
        WHERE import_id = %s
    """, (json.dumps(errors), rows_rejected, import_id))
    db_conn.commit()
    cursor.close()


def error_response(message: str, details: Any = None):
    """Standard error response"""
    return {
        'statusCode': 400,
        'body': json.dumps({
            'success': False,
            'error': message,
            'details': details
        })
    }

